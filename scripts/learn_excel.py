from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook


def read_excel(excel_full_name):
    wb = load_workbook(excel_full_name)
    sheets = wb.sheetnames
    sheet_first = sheets[0]
    ws = wb[sheet_first]
    rows = ws.rows

    new_rows = []
    for row in rows:
        line = [col.value for col in row]
        col_list = str(line[-1]).replace('\n', '').split(',')    # 简单的数据清洗以归整数据
        for v in col_list:
            new_rows.append([line[0], line[1], line[2], line[3], v])
    return new_rows


def write_excel(data_list, w2file):
    wb = Workbook()
    ws = wb.active
    table_title = ['SOURCE_TABLE', 'TARGET_TABLE', 'DATA_SRC', 'TARGET_SCHEMA', 'FIELDS']
    ws.append(table_title)
    for v in data_list:
        ws.append(v)
    wb.save(w2file)


if __name__ == '__main__':
    input_excel_name = r'D:\WeChat Files\whk11061023\FileStorage\File\2019-08/DIM_ETL_CONFIG.xlsx'
    output_excel_name = r'D:\WeChat Files\whk11061023\FileStorage\File\2019-08/DIM_ETL_CONFIG_new1.xlsx'
    write_excel(read_excel(input_excel_name), output_excel_name)
    exit()
