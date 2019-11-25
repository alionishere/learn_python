#!/usr/bin/python
# -*- coding = utf-8 -*-
import sys
import numpy
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook


# str = 'aa' + 'b' + \
#       'c'
# print(str)
# print(sys.path[0])


def read(excel_full_name):
    # 依据excel文件的全路径名获取整个excel对象
    wb = load_workbook(excel_full_name)
    # 获取所有sheet对象名
    sheets = wb.sheetnames
    # 获取第一个sheet页名
    fst_sheet = sheets[0]
    # 依据sheet名获取该页数据对象
    ws = wb[fst_sheet]
    # 通过rows属性，获取所有行对象
    rows = ws.rows
    # 遍历行对象
    for row in rows:
        # 遍历row，获取每行数据
        # for col in row:  # col: 每个单元格对象
        #     print(col.value)  # value:  获取每个单元格值
        #     print('*'*50)
        line = [col.value for col in row]  # 将单行数据组合为一个列表
        print(line)


def write(data, w2file):
    # 创建一个excel对象
    wb = Workbook()
    # 获取一个excel sheet页
    ws = wb.active
    # 向sheet页添加数据，默认是顺序添加
    table_title = [1, 2, 3, 4, 5, 6]
    ws.append(table_title)
    # 写入目标文件
    wb.save(w2file)


if __name__ == '__main__':
    input_excel_name = r'D:\WeChat Files\whk11061023\FileStorage\File\2019-08/DIM_ETL_CONFIG.xlsx'
    excel_file = r'E:\test_read.xlsx'
    read(excel_file)
