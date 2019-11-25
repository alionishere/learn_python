from openpyxl import Workbook
import datetime
import time

wb = Workbook()
ws = wb.active

table_title = ['SOURCE_TABLE', 'TARGET_TABLE', 'DATA_SRC', 'TARGET_SCHEMA', 'FIELDS']
ws.append(table_title)
ws['A1'] = 'Hello'
ws['B2'] = '你好'
ws['A2'] = datetime.datetime.now()
ws['B1'] = time.strftime("%Y%m-%d %H:%M:%S", time.localtime())

wb.save(r'e:/sample.xlsx')
